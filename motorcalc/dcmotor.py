import numpy as np
import matplotlib.pyplot as plt
import datetime
from numpy.core.arrayprint import _none_or_positive_arg
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from texttable import Texttable
from dataclasses import dataclass
from typing import List

APP_NAME = "dcmotor.py"
APP_VERSION = "1.0"

def omega_to_speed_rpm(omega: np.array) -> np.array:
    """convert angular speed in rad/s to rpm"""
    return omega/np.pi*30

def speed_rpm_to_omega(speed_rpm: np.array) -> np.array:
    """convert rpm to angular speed in rad/s"""
    return speed_rpm*np.pi/30

def angle_to_number_of_rotations(angle: np.array) -> np.array:
    """Convert an angular position in rad to number of rotations"""
    return angle/(2*np.pi)

def number_of_rotations_to_angle(number_rot: np.array, is_absolute: bool = True) -> np.array:
    """Convert a given number of rotations to an angular positition in rad"""
    angle = number_rot * 2 * np.pi
    if is_absolute:
        return angle
    return np.mod(angle, 2*np.pi)

@dataclass(slots=True)
class CGearbox:
    """class to represent a gearbox on a DC motor"""
    ratio: float            # gearbox ratio of output speed over input speed
    efficiency: float       # ratio of mechanical output power over mechanical input power
    name: str               # name of the gearbox


        
class CDCMotor():
    """
    A class used to represent a DC Motor for calculation
    ...

    Attributes
    ----------
    U_N : float
        nominal Voltage in V
    I_0 : float
        noload current in A
    k_M : float
        torque constant in Nm/A
    R : float
        terminal resistance in Ohm
    H : float
        terminal inductance in mH (for information only)
    Theta : float
        rotor moment of inertia in gcm^2 (for information only)
    nPoints : int
        number of points to be plotted in graph
    n_WP : float
        required speed at working point
    M_WP : float
        required torque at working point
    motor_name : str
        name of the motor used for graph plotting
    application : str
        short description of application for graph plotting
    file_name : str
        name of Excel file for report generation
    

    Methods
    -------
    calc_motor_values()
        Calculates the motor performance curves for a given set of parameter
    """
    def __init__(self, U_N:float=0.0, I_0:float=0.0, k_M:float=0.0, R:float=0.0, H:float=0.0, Theta:float=0.0, \
        nPoints:int=100, n_WP:float=0.0, M_WP:float=0.0, motor_name:str='', application:str='', file_name:str='output.xlsx'):
        """
        Parameters
        ----------
        U_N : float
            nominal Voltage in V
        I_0 : float
            noload current in A
        k_M : float
            torque constant in Nm/A
        R : float
            terminal resistance in Ohm
        H : float
            terminal inductance in mH (for information only)
        Theta : float
            rotor moment of inertia in gcm^2 (for information only)
        nPoints : int
            number of points to be plotted in graph
        n_WP : float
            required speed at working point
        M_WP : float
            required torque at working point
        motor_name : str
            name of the motor used for graph plotting
        application : str
            short description of application for graph plotting
        file_name : str
        name of Excel file for report generation
        """
        self.U_N = U_N                      # nominal Voltage in V
        self.I_0 = I_0                      # no load current in A
        self.k_M = k_M                      # torque constant in Nm/A
        self.R = R                          # terminal resistance in Ohms
        self.H = H                          # terminal inductance in mH
        self.Theta = Theta                  # rotor inertia in gcm^2 
        self.nPoints = nPoints              # number of points to be plotted in graph
        self.n_WP = n_WP                    # required speed at working point
        self.M_WP = M_WP                    # required torque at working point
        self.motor_name = motor_name        # motor name for text field in plot
        self.application = application      # name of application for title
        self.file_name = file_name          # name of file for excel export
        self.gearboxes: List[CGearbox] = [] # Gearbox stages to be added with add_gearbox
        self.calc_scalar_parameter()        # calculates all other scalar motor parameter

    
    def add_gearbox(self, gb: CGearbox):
        """add a gearbox class to the gearboxes-list"""
        self.gearboxes.append(gb)
        self.calc_scalar_parameter()

    def calc_gearboxes_ratio(self) -> float:
        """calculate the effective ratio of the gearboxes in the gearboxes list"""
        ratio = 1.0
        for gb in self.gearboxes:
            ratio*=gb.ratio
        return ratio

    def calc_gearboxes_efficiency(self) -> float:
        """calculate the effective efficiency in the gearboxes list"""
        efficiency = 1.0
        for gb in self.gearboxes:
            efficiency*=gb.efficiency
        return efficiency
    
    def calc_I_from_M(self, M:np.array)->np.array:
        """
        Calculates the current at a given torque value
        
        Parameters:
        -----------
        M : numpy.array
            Array of torque values
        """
        return (M+self.M_0)/self.k_M

    def calc_M_0(self)->float:
        """Calculates the no-load torque"""
        return self.k_M*self.I_0

    def calc_I_S(self)->float:
        """Calculates the stall current"""
        return self.U_N/self.R

    def calc_M_S(self)->float:
        """Calculates the stall torque"""
        return self.k_M*(self.I_S-self.I_0)

    def calc_n_from_M(self, M:np.array)->np.array:
        """
        Calculates the speed for a given set of torque values
        
        Parameters:
        -----------
        M : numpy array
            Array of torque values
        """
        omega=self.b+self.a*(M+self.M_0)
        return omega_to_speed_rpm(omega) 

    def calc_M_from_omega(self, omega:np.array)->np.array:
        """
        Calculates torque values for a given set of speed values
        
        Parameters:
        -----------
        omega : numpy array
            Array of angular speed values
        """
        M = (omega-self.b)/self.a-self.M_0
        return M


    def calc_M_from_n(self, n:np.array)->np.array:
        """
        Calculates torque values for a given set of speed values
        
        Parameters:
        -----------
        n : numpy array
            Array of speed values
        """
        omega=speed_rpm_to_omega(n)
        return self.calc_M_from_omega(omega=omega)


    def calc_P_el_from_M(self, M:np.array)->np.array:
        """
        Calculates electrical power values for a given set of torque values
        
        Parameters:
        -----------
        M : numpy array
            Array of torque values
        """
        return self.b*(M+self.M_0)

    def calc_P_mech_from_M(self, M:np.array)->np.array:
        """
        Calculates mechanical power values for a given set of torque values
        
        Parameters:
        -----------
        M : numpy array
            Array of torque values
        """
        return M*self.b+self.a*M**2+self.a*M*self.M_0

    def calc_eta_from_M(self, M:np.array)->np.array:
        """
        Calculates efficiency values for a given set of torque values
        
        Parameters:
        -----------
        M : numpy array
            Array of torque values
        """
        c = 1/(M+self.M_0)
        frac=self.a/self.b
        return c*(M+frac*M**2+frac*M*self.M_0)


    def calc_scalar_parameter(self):
        """
        Calculates all scalar motor parameter, to be stored in the class object.
        """
        self.b = self.U_N/self.k_M
        self.a = -self.R/self.k_M**2

        ## no load torque (Reibmoment) in Nm
        self.M_0 = self.calc_M_0()
       
        ## stall current in A
        self.I_S = self.calc_I_S()

        ## stall torque in Nm
        self.M_S = self.calc_M_S()

        ## no-load speed
        self.n_0 = self.calc_n_from_M(0.0)

        ## torque @ max efficiency
        self.M_meff = self.M_0*(np.sqrt(self.M_S/self.M_0+1)-1)

        ## current @ max efficiency
        self.I_meff = np.sqrt(self.M_S*self.M_0+self.M_0**2)/self.k_M
                
        ## max efficiency
        self.eta_max = (1-np.sqrt(self.I_0/self.I_S))**2

        ## power @ max efficiency
        self.P_meff=self.a*self.M_0*self.M_S+self.a*self.M_0**2-self.b*self.M_0 \
            + np.sqrt(self.M_S*self.M_0+self.M_0**2) * (self.b-self.a*self.M_0) \

        ## max power
        self.M_maxpower = 0.5*self.M_S
        self.P_maxpower = self.a/4*self.M_S**2+self.a/2*self.M_S*self.M_0+self.b/2*self.M_S

        ## load speed (speed @ max efficiency)
        self.n_meff = (self.b+self.a*np.sqrt(self.M_S*self.M_0+self.M_0**2))*30/np.pi

        self.calc_scalar_parameter_gb()


    def estimate_delta_I_from_gb_efficiency(self) -> float:
        """Estimate the additive current introduced by the gearboxes"""
        gb_efficiency = self.calc_gearboxes_efficiency()
        delta_M = self.M_meff*(1-gb_efficiency)
        k_M = self.k_M
        return delta_M/k_M

    def calc_scalar_parameter_gb(self):
        """
        Calculates all scalar motor parameter including the gearboxes assuming a constant loss moment 
        estimated from the efficiency of the gear boxes. Relies on calc_scalar_parameter to be executed beforehand.
        Estimation of the loss moment is empiric, based on the moment of highest efficiency. Results underestimate
        the losses due to efficiency. 
        """
        if self.gearboxes==[]:
            return
        
        gb_ratio = self.calc_gearboxes_ratio()
        # print(gb_ratio)

        delta_I = self.estimate_delta_I_from_gb_efficiency()
        # print(delta_I)

        self.k_M /= gb_ratio
        self.I_0 += delta_I

        self.b = self.U_N/self.k_M
        self.a = -self.R/self.k_M**2

        ## no load torque (Reibmoment) in Nm
        self.M_0 = self.calc_M_0()
       
        ## stall current in A
        self.I_S = self.calc_I_S()

        ## stall torque in Nm
        self.M_S = self.calc_M_S()

        ## no-load speed
        self.n_0 = self.calc_n_from_M(0.0)

        ## torque @ max efficiency
        self.M_meff = self.M_0*(np.sqrt(self.M_S/self.M_0+1)-1)

        ## current @ max efficiency
        self.I_meff = np.sqrt(self.M_S*self.M_0+self.M_0**2)/self.k_M
                
        ## max efficiency
        self.eta_max = (1-np.sqrt(self.I_0/self.I_S))**2

        ## power @ max efficiency
        self.P_meff=self.a*self.M_0*self.M_S+self.a*self.M_0**2-self.b*self.M_0 \
            + np.sqrt(self.M_S*self.M_0+self.M_0**2) * (self.b-self.a*self.M_0) \

        ## max power
        self.M_maxpower = 0.5*self.M_S
        self.P_maxpower = self.a/4*self.M_S**2+self.a/2*self.M_S*self.M_0+self.b/2*self.M_S

        ## load speed (speed @ max efficiency)
        self.n_meff = (self.b+self.a*np.sqrt(self.M_S*self.M_0+self.M_0**2))*30/np.pi


    def calc_torque_from_current(self, I:np.array)->np.array:
        """
        Calculates torque values for a given set of current values
        
        Parameters:
        -----------
        M : numpy array
            Array of current values
        """
        return I*self.k_M 

    def tune_voltage_to_working_point(self):
        """
        Set the member variable value U_N to a value where the given member variable values n_WP (working point speed)
        and M_WP (working point torque) will be met. 
        """
        self.U_N=self.R*(self.M_WP+self.M_0)/self.k_M+self.k_M*self.n_WP*np.pi/30.0
        self.calc_scalar_parameter()

    def calc_performance_curves(self):
        ## torque range available (used for plotting)
        self.M = np.linspace(0.0,self.M_S,self.nPoints,endpoint=True)
        ## current curve over torque range
        self.I = self.calc_I_from_M(self.M)
        ## speed over torque
        self.n = self.calc_n_from_M(self.M)
        ## electrical power over torque range
        self.P_el = self.calc_P_el_from_M(self.M)
        ## mechanical power over torque range
        self.P_mech = self.calc_P_mech_from_M(self.M)
        ## efficiency
        self.eta = self.P_mech / self.P_el

    def parameter_txt(self) -> str:
        """
        Generates a string with motor parameter text
        """
        out_str='\n\n'
        out_str+='INPUT PARAMETER:\n'
        out_str+=self._dash_line()
        out_str+='parameter\tvoltage\t\tterm. resist.\tno-load cur.\tno-load speed\ttorque const.\n'
        out_str+=self._dash_line()
        out_str+='unit\t\tVolt\t\tOhm\t\tAmpere\t\tRPM\t\tNm/A\n'
        out_str+='value\t\t{:0.1f}\t\t{:0.2f}\t\t{:0.3f}\t\t{:0.0f}\t\t{:0.3f}\n\n'.format(self.U_N,self.R,self.I_0,self.n_0,self.k_M)
        out_str+='PERFORMANCE DATA:\n'
        out_str+=self._dash_line()
        out_str+='parameter\tunit\tno-load\t\t@max eff.\t@max power\tstall\t\t@working point\n'
        out_str+=self._dash_line()
        out_str+='speed\t\tRPM\t{:0.0f}\t\t{:0.0f}\t\t{:0.0f}\t\t{:0.0f}\t\t{:0.0f}\n'.format( \
            self.n_0, self.n_meff, self.calc_n_from_M(self.M_maxpower), 0, self.calc_n_from_M(self.M_WP))
        out_str+='current\t\tA\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\n'.format( \
            self.I_0, self.I_meff, self.calc_I_from_M(self.M_maxpower), self.I_S, self.calc_I_from_M(self.M_WP))
        out_str+='torque\t\tNm\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\n'.format( \
            self.M_0, self.M_meff, self.M_maxpower, self.M_S, self.M_WP)
        out_str+='power\t\tW\t{:0.2f}\t\t{:0.2f}\t\t{:0.2f}\t\t{:0.2f}\t\t{:0.2f}\n'.format( \
            0, self.P_meff, self.P_maxpower, 0, self.calc_P_mech_from_M(self.M_WP))
        out_str+='eff.\t\t%\t{:0.1f}\t\t{:0.1f}\t\t{:0.1f}\t\t{:0.1f}\t\t{:0.1f}\n\n'.format(0, self.eta_max*100.0, \
            self.calc_eta_from_M(self.M_maxpower)*100.0, 0.0, self.calc_eta_from_M(self.M_WP)*100.0)
        return out_str

    def _dash_line(self, length: int = 102, linebreak: bool = True):
        dash_str=''
        for _ in range(length):
            dash_str += '-'
        if not linebreak:
            return dash_str
        return dash_str + '\n'

    def __str__(self):
        return self.parameter_txt()


    def list_spec_table(self):
        """
        Prints a specification table to the command line
        """
        t = Texttable()
        t.add_row(["No","Parameter","Unit","Value"])
        t.add_row([1,"Voltage","V",self.U_N])
        t.add_row([2,"Terminal resistance","Ω",self.R])
        t.add_row([3, "Terminal inductance", "mH", self.H])
        t.add_row([4, "No-load speed", "rpm", int(self.n_0)])
        t.add_row([5, "No-load current", "A", self.I_0])
        t.add_row([6, "Nominal torque", "mNm", 1000.0*self.M_WP])
        t.add_row([7, "Nominal speed", "rpm", int(self.calc_n_from_M(self.M_WP))])
        t.add_row([8, "Nominal current", "A", self.calc_I_from_M(self.M_WP)])
        t.add_row([9, "Max. output power", "W", self.P_maxpower])        
        t.add_row([10, "Max. efficiency", "%", self.eta_max*100.0])
        t.add_row([11, "Back-EMF constant", "mV/rpm", self.k_M/9.81*1000.0])
        t.add_row([12, "Torque constant", "mNm/A", 1000*self.k_M])
        t.add_row([13, "Speed/torque gradient", "rpm/mNm", self.n_0/self.M_S/1000.0])
        t.add_row([14, "Rotor inertia", "gcm^2", self.Theta])
        # 15 Weight g 21
        # 16 Thermal resistance housing-ambient K/W 8
        # 17 Thermal resistance winding-housing K/W 9.5
        # 18 Thermal time constant motor s 354
        # 19 Thermal time constant winding s 23
        # 20 Operating temperature range °C -40 ~ +120
        # 21 Thermal class of winding °C 155
        # 22 Axial play mm 0.012
        # 23 Radial play mm 0.08
        # 24 Axial load dynamic N 1.5
        # 25 Axial load static N 37
        # 26 Radial load at 3 mm from mounting face N 12
        # 27 No. of pole pairs 4
        # 28 Bearings 2 ball bearings
        # 29 Commutation Sensorless
        # 30 Protection class IP 20
        print(t.draw())


    def export_to_excel(self):
        """
        Exports the results to excel. Uses the file name given in the member variable self.file_name
        """
        exWB = Workbook()
        ws = exWB.active
        try:
            if self.motor_name != '':
                ws.title = self.motor_name
        except:
            pass
        ws.append(["DC Motor Calculation: " + self.motor_name])
        ws['A1'].font=Font(size=16,bold=True)
        ws.append(["File generated on:", "", datetime.datetime.now().strftime("%d-%b-%y"), datetime.datetime.now().strftime("%H:%M:%S")])
        ws['A2'].font=Font(size=12,bold=True,italic=True)
        ws['B2'].font=Font(size=12,bold=True,italic=True)
        ws.append(["File generated by: " + APP_NAME + ", Version " + APP_VERSION])
        ws['A3'].font=Font(size=12,bold=True,italic=True)
        ws.append([])

        row=5
        col=1
        titleStr = "Input data DC motor"
        headerStr = ['', 'voltage', 'term. resist.', 'no-load cur.', 'no-load speed', 'torque const.']
        data = [ \
            ['unit', 'Volt', 'Ohm', 'A', 'RPM', 'mNm/A'], \
            ['value', self.U_N, self.R, self.I_0, self.n_0, self.k_M] \
            ]
        row, col = self._excel_add_table_to_worksheet(ws, row, col, titleStr, headerStr, data)
        
        row += 1
        col = 1
        titleStr = "DC motor performance data"
        headerStr = ['', 'unit', '@no-load', '@max eff.', '@max power', 'stall']
        data = [ \
            ['speed', 'RPM', self.n_0, self.n_meff, self.calc_n_from_M(self.M_maxpower),0], \
            ['current', 'A', self.I_0, self.I_meff, self.calc_I_from_M(self.M_maxpower),self.I_S], \
            ['torque', 'Nm', self.M_0, self.M_meff, self.M_maxpower,self.M_S], \
            ['power', 'W', 0, self.P_meff, self.P_maxpower,0], \
            ['eff.', '%', 0, self.eta_max*100.0, self.calc_eta_from_M(self.M_maxpower)*100.0,0.0] \
            ]
        row, col = self._excel_add_table_to_worksheet(ws, row, col, titleStr, headerStr, data)
        if self.file_name == "":
            if not self.motor_name:
                self.file_name = "MotorCalc_output.xlsx"
            else:
                self.file_name = self.motor_name + ".xlsx"
        exWB.save(filename=self.file_name)
        return(row, col)
    
    def _excel_add_table_to_worksheet(self, ws, row, col, titleStr, headerStr, data):
        ws.cell(row, col, value = titleStr).font = Font(size=12,bold=True)
        row += 1
        for hstr in headerStr:
            ws.cell(row,col,value=hstr).font=Font(bold=True)
            col+=1
        row+=1
        col=1

        for dpoints in data:
            for dp in dpoints:
                ws.cell(row,col,value=dp)
                if col == 1:
                    ws.cell(row,col).font=Font(size=12, bold=True)

                col+=1
            row += 1
            col = 1
        return row,col 


def Main():
    # R=1.44 (-25°C)
    # R=2.00 (125°C)
    
    c = 4.7
    U_N = 10
    T = -40
    m_T = (2.0-1.44)/(125+25)
    T_0 = 1.44 - m_T*(-25)
    R = m_T*T+T_0*c
    k_M = 0.005985*np.sqrt(c)
    I_0 = 0.13*6/U_N
    n_WP = 300
    M_WP = 0.005
    dcmotor=CDCMotor(U_N=U_N, I_0=I_0, k_M=k_M, R=R, H=0.61, Theta=6.7, n_WP=n_WP, M_WP=M_WP, application="166_A / LiDAR", motor_name="BO2015_Version 10V")
    print(dcmotor)
    dcmotor.tune_voltage_to_working_point()
    dcmotor.list_spec_table()
 
if __name__ == "__main__":
    Main()