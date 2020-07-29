## Motorcalculation Software
## Gerrit Kocherscheidt, KOCO automotive GmbH
## Date 28-Jul-2020

import numpy as np
import matplotlib.pyplot as plt
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from texttable import Texttable

APP_NAME = "MotorCalc.py"
APP_VERSION = "0.3"

class CDCMotor :
    """
    A class used to represent a DC Motor for calculation

    ...

    Attributes
    ----------
    U_N : float
        nominal Voltage in V
    I_0 : float
        noload current in A
    k_M : float
        torque constant in Nm/A
    R : float
        terminal resistance in Ohm
    H : float
        terminal inductance in mH (for information only)
    Theta : float
        rotor moment of inertia in gcm^2 (for information only)
    nPoints : int
        number of points to be plotted in graph
    n_WP : float
        required speed at working point
    M_WP : float
        required torque at working point
    motor_name : str
        name of the motor used for graph plotting
    application : str
        short description of application for graph plotting
    file_name : str
        name of Excel file for report generation
    

    Methods
    -------
    calc_motor_values()
        Calculates the motor performance curves for a given set of parameter
    """
    def __init__(self, U_N:float=0.0, I_0:float=0.0, k_M:float=0.0, R:float=0.0, H:float=0.0, Theta:float=0.0, \
        nPoints:int=100, n_WP:float=0.0, M_WP:float=0.0, motor_name:str='', application:str='', file_name:str='output.xlsx'):
        """
        Parameters
        ----------
        U_N : float
            nominal Voltage in V
        I_0 : float
            noload current in A
        k_M : float
            torque constant in Nm/A
        R : float
            terminal resistance in Ohm
        H : float
            terminal inductance in mH (for information only)
        Theta : float
            rotor moment of inertia in gcm^2 (for information only)
        nPoints : int
            number of points to be plotted in graph
        n_WP : float
            required speed at working point
        M_WP : float
            required torque at working point
        motor_name : str
            name of the motor used for graph plotting
        application : str
            short description of application for graph plotting
        file_name : str
        name of Excel file for report generation
        """
        self.U_N = U_N                  # nominal Voltage in V
        self.I_0 = I_0                  # no load current in A
        self.k_M = k_M                  # torque constant in Nm/A
        self.R = R                      # terminal resistance in Ohms
        self.H = H                      # terminal inductance in mH
        self.Theta = Theta              # rotor inertia in gcm^2 
        self.nPoints = nPoints          # number of points to be plotted in graph
        self.n_WP = n_WP                # required speed at working point
        self.M_WP = M_WP                # required torque at working point
        self.motor_name = motor_name    # motor name for text field in plot
        self.application = application  # name of application for title
        self.file_name = file_name      # name of file for excel export
        self.calc_motor_values()

    def calc_I_from_M(self, M:np.array)->np.array:
        """
        Calculates the current at a given torque value
        
        Parameters:
        -----------
        M : numpy.array
            Array of torque values
        """
        return((M+self.M_0)/self.k_M)

    def calc_M_0(self)->float:
        """Calculates the no-load torque"""
        return(self.k_M*self.I_0)

    def calc_I_S(self)->float:
        """Calculates the stall current"""
        return(self.U_N/self.R)

    def calc_M_S(self)->float:
        """Calculates the stall torque"""
        return(self.k_M*(self.I_S-self.I_0))

    def calc_n_from_M(self, M:np.array)->np.array:
        """
        Calculates the speed for a given set of torque values
        
        Parameters:
        -----------
        M : numpy array
            Array of torque values
        """
        omega=self.b+self.a*(M+self.M_0)
        return(omega*30/np.pi)

    def calc_M_from_n(self, n:np.array)->np.array:
        """
        Calculates torque values for a given set of speed values
        
        Parameters:
        -----------
        n : numpy array
            Array of speed values
        """
        omega=n*np.pi/30
        return((omega-self.b)/self.a-self.M_0)

    def calc_P_el_from_M(self, M:np.array)->np.array:
        """
        Calculates electrical power values for a given set of torque values
        
        Parameters:
        -----------
        M : numpy array
            Array of torque values
        """
        return(self.b*(M+self.M_0))

    def calc_P_mech_from_M(self, M:np.array)->np.array:
        """
        Calculates mechanical power values for a given set of torque values
        
        Parameters:
        -----------
        M : numpy array
            Array of torque values
        """
        return(M*self.b+self.a*M**2+self.a*M*self.M_0)

    def calc_eta_from_M(self, M:np.array)->np.array:
        """
        Calculates efficiency values for a given set of torque values
        
        Parameters:
        -----------
        M : numpy array
            Array of torque values
        """
        c = 1/(M+self.M_0)
        frac=self.a/self.b
        return(c*(M+frac*M**2+frac*M*self.M_0))


    def calc_motor_values(self):
        """
        Calculates all values of the motor modell
        """
        self.b = self.U_N/self.k_M
        self.a = -self.R/self.k_M**2

        ## no load torque (Reibmoment) in Nm
        self.M_0 = self.calc_M_0()
       
        ## stall current in A
        self.I_S = self.calc_I_S()

        ## stall torque in Nm
        self.M_S = self.calc_M_S()

        ## torque range available (used for plotting)
        self.M = np.linspace(0.0,self.M_S,self.nPoints,endpoint=True)

        ## current curve over torque range
        self.I = self.calc_I_from_M(self.M)

        ## speed curve over torque range
##        self.omega = (self.U_N-self.R*self.I)/self.k_M
##
##        self.n = self.omega*30/np.pi
        self.n = self.calc_n_from_M(self.M)
        self.n_0 = self.n[0]

        ## electrical power over torque range
        self.P_el = self.calc_P_el_from_M(self.M)

        ## mechanical power over torque range
        self.P_mech = self.calc_P_mech_from_M(self.M)

        ## efficiency
        self.eta = self.P_mech / self.P_el

        ## torque @ max efficiency
        self.M_meff = self.M_0*(np.sqrt(self.M_S/self.M_0+1)-1)

        ## current @ max efficiency
        self.I_meff = np.sqrt(self.M_S*self.M_0+self.M_0**2)/self.k_M
                
        ## max efficiency
        self.eta_max = (1-np.sqrt(self.I_0/self.I_S))**2

        ## power @ max efficiency
        self.P_meff=self.a*self.M_0*self.M_S+self.a*self.M_0**2-self.b*self.M_0 + np.sqrt(self.M_S*self.M_0+self.M_0**2) * (self.b-self.a*self.M_0)

        ## max power
        self.M_maxpower = 0.5*self.M_S
        self.P_maxpower = self.a/4*self.M_S**2+self.a/2*self.M_S*self.M_0+self.b/2*self.M_S

        ## load speed (speed @ max efficiency)
        self.n_meff = (self.b+self.a*np.sqrt(self.M_S*self.M_0+self.M_0**2))*30/np.pi

    def calc_torque_from_current(self, I:np.array)->np.array:
        """
        Calculates torque values for a given set of current values
        
        Parameters:
        -----------
        M : numpy array
            Array of current values
        """
        return(I*self.k_M)

    def tune_voltage_to_working_point(self):
        """
        Set the member variable value U_N to a value where the given member variable values n_WP (working point speed)
        and M_WP (working point torque) will be met. 
        """
        self.U_N=self.R*(self.M_WP+self.M_0)/self.k_M+self.k_M*self.n_WP*np.pi/30.0
        self.calc_motor_values()


    def print_parameter(self):
        """
        Prints a set of system values to the command line
        """
        print('input parameter')
        print('parameter\tvoltage\t\tterm. resist.\tno-load cur.\tno-load speed\ttorque const.')
        print('unit\t\tVolt\t\tOhm\t\tAmpere\t\tRPM\t\tNm/A')
        print('value\t\t{:0.1f}\t\t{:0.2f}\t\t{:0.3f}\t\t{:0.0f}\t\t{:0.3f}'.format(self.U_N,self.R,self.I_0,self.n_0,self.k_M))
        print('')
        print('motor performance data:')
        print('parameter\tunit\tno-load\t\t@max eff.\t@max power\tstall\t@working point')
        print('speed\t\tRPM\t{:0.0f}\t\t{:0.0f}\t\t{:0.0f}\t\t{:0.0f}\t\t{:0.0f}'.format( \
            self.n_0, self.n_meff, self.calc_n_from_M(self.M_maxpower), 0, self.calc_n_from_M(self.M_WP)))
        print('current\t\tA\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}'.format( \
            self.I_0, self.I_meff, self.calc_I_from_M(self.M_maxpower), self.I_S, self.calc_I_from_M(self.M_WP)))
        print('torque\t\tNm\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}'.format( \
            self.M_0, self.M_meff, self.M_maxpower, self.M_S, self.M_WP))
        print('power\t\tW\t{:0.2f}\t\t{:0.2f}\t\t{:0.2f}\t\t{:0.2f}\t\t{:0.2f}'.format( \
            0, self.P_meff, self.P_maxpower, 0, self.calc_P_mech_from_M(self.M_WP)))
        print('eff.\t\t%\t{:0.1f}\t\t{:0.1f}\t\t{:0.1f}\t\t{:0.1f}\t\t{:0.1f}'.format(0, self.eta_max*100.0, \
            self.calc_eta_from_M(self.M_maxpower)*100.0, 0.0, self.calc_eta_from_M(self.M_WP)*100.0))
        print('')

    def list_spec_table(self):
        """
        Prints a specification table to the command line
        """
        t = Texttable()
        t.add_row(["No","Parameter","Unit","Value"])
        t.add_row([1,"Voltage","V",self.U_N])
        t.add_row([2,"Terminal resistance","Ω",self.R])
        t.add_row([3, "Terminal inductance", "mH", self.H])
        t.add_row([4, "No-load speed", "rpm", int(self.n_0)])
        t.add_row([5, "No-load current", "A", self.I_0])
        t.add_row([6, "Nominal torque", "mNm", 1000.0*self.M_WP])
        t.add_row([7, "Nominal speed", "rpm", int(self.calc_n_from_M(self.M_WP))])
        t.add_row([8, "Nominal current", "A", self.calc_I_from_M(self.M_WP)])
        t.add_row([9, "Max. output power", "W", self.P_maxpower])        
        t.add_row([10, "Max. efficiency", "%", self.eta_max])
        t.add_row([11, "Back-EMF constant", "mV/rpm", self.k_M/9.81*1000.0])
        t.add_row([12, "Torque constant", "mNm/A", 1000*self.k_M])
        t.add_row([13, "Speed/torque gradient", "rpm/mNm", self.n_0/self.M_S/1000.0])
        t.add_row([14, "Rotor inertia", "gcm^2", self.Theta])
        # 15 Weight g 21
        # 16 Thermal resistance housing-ambient K/W 8
        # 17 Thermal resistance winding-housing K/W 9.5
        # 18 Thermal time constant motor s 354
        # 19 Thermal time constant winding s 23
        # 20 Operating temperature range °C -40 ~ +120
        # 21 Thermal class of winding °C 155
        # 22 Axial play mm 0.012
        # 23 Radial play mm 0.08
        # 24 Axial load dynamic N 1.5
        # 25 Axial load static N 37
        # 26 Radial load at 3 mm from mounting face N 12
        # 27 No. of pole pairs 4
        # 28 Bearings 2 ball bearings
        # 29 Commutation Sensorless
        # 30 Protection class IP 20
        print(t.draw())


    def export_to_excel(self):
        """
        Exports the results to excel. Uses the file name given in the member variable self.file_name
        """
        exWB = Workbook()
        ws = exWB.active
        try:
            if self.motor_name != '':
                ws.title = self.motor_name
        except:
            pass
        ws.append(["DC Motor Calculation: " + self.motor_name])
        ws['A1'].font=Font(size=16,bold=True)
        ws.append(["File generated on:", "", datetime.datetime.now().strftime("%d-%b-%y"), datetime.datetime.now().strftime("%H:%M:%S")])
        ws['A2'].font=Font(size=12,bold=True,italic=True)
        ws['B2'].font=Font(size=12,bold=True,italic=True)
        ws.append(["File generated by: " + APP_NAME + ", Version " + APP_VERSION])
        ws['A3'].font=Font(size=12,bold=True,italic=True)
        ws.append([])

        row=5
        col=1
        titleStr = "Input data DC motor"
        headerStr = ['', 'voltage', 'term. resist.', 'no-load cur.', 'no-load speed', 'torque const.']
        data = [ \
            ['unit', 'Volt', 'Ohm', 'A', 'RPM', 'mNm/A'], \
            ['value', self.U_N, self.R, self.I_0, self.n_0, self.k_M] \
            ]
        row, col = self._excel_add_table_to_worksheet(ws, row, col, titleStr, headerStr, data)
        
        row += 1
        col = 1
        titleStr = "DC motor performance data"
        headerStr = ['', 'unit', '@no-load', '@max eff.', '@max power', 'stall']
        data = [ \
            ['speed', 'RPM', self.n_0, self.n_meff, self.calc_n_from_M(self.M_maxpower),0], \
            ['current', 'A', self.I_0, self.I_meff, self.calc_I_from_M(self.M_maxpower),self.I_S], \
            ['torque', 'Nm', self.M_0, self.M_meff, self.M_maxpower,self.M_S], \
            ['power', 'W', 0, self.P_meff, self.P_maxpower,0], \
            ['eff.', '%', 0, self.eta_max*100.0, self.calc_eta_from_M(self.M_maxpower)*100.0,0.0] \
            ]
        row, col = self._excel_add_table_to_worksheet(ws, row, col, titleStr, headerStr, data)
        if self.file_name == "":
            if not self.motor_name:
                self.file_name = "MotorCalc_output.xlsx"
            else:
                self.file_name = self.motor_name + ".xlsx"
        exWB.save(filename=self.file_name)
        return(row, col)
    
    def _excel_add_table_to_worksheet(self, ws, row, col, titleStr, headerStr, data):
        ws.cell(row, col, value = titleStr).font = Font(size=12,bold=True)
        row += 1
        for hstr in headerStr:
            ws.cell(row,col,value=hstr).font=Font(bold=True)
            col+=1
        row+=1
        col=1

        for dpoints in data:
            for dp in dpoints:
                ws.cell(row,col,value=dp)
                if col == 1:
                    ws.cell(row,col).font=Font(size=12, bold=True)

                col+=1
            row += 1
            col = 1
        return(row,col)


    def plotCurves(self, addVoltagesSpeed:list=None):
        """
        Plots system values into a matplotlib graph

        Parameters:
        -----------
        addVoltagesSpeed : list of floats
            List of additional voltages that are used to plot additional n-over-M-curves
        """
        fig=plt.figure(figsize=(12,8))
        fig.patch.set_facecolor('white')
        host = plt.subplot(111)
        host.set_title(self.application)
        host.patch.set_facecolor('white')

        #generate host plot
        plt.subplots_adjust(right=0.75)
        plt.subplots_adjust(left=0.1)
        plt.subplots_adjust(bottom=0.10)
        host.plot(self.M*1000.0,self.I,color="red")
        host.plot([1000.0*self.M_meff,1000.0*self.M_meff],[0.0, 1.2*self.I_S],":",color="black")
        host.plot([1000.0*self.M_maxpower,1000.0*self.M_maxpower],[0.0, 1.2*self.I_S],":",color="black")
        if self.M_WP != 0:
            host.plot([1000.0*self.M_WP,1000.0*self.M_WP],[0.0, 1.2*self.I_S],"-",color="cyan")
        host.set_xlabel("torque (mNm)")
        host.set_ylabel("current (A)")
        host.yaxis.label.set_color("red")
        host.spines["left"].set_color("red")
        host.tick_params(axis="y", colors="red")
        host.grid(True)
        host.set_xlim(0,1000.0*self.M_S)
        host.set_ylim(0,1.2*self.I_S)
        

        ax_power=host.twinx()
        ax_power.patch.set_facecolor('white')
        ax_power.plot(self.M*1000.0,self.P_mech,".-",color="black")
        ax_power.plot(1000.0*self.M_meff,self.P_meff,"d",color="red")
        ax_power.plot(1000.0*self.M_maxpower,self.P_maxpower,"d",color="red")
        if self.M_WP != 0 and self.n_WP != 0:
            P_mech_WP = self.M_WP * self.n_WP * np.pi / 30.0
            ax_power.plot(1000.0*self.M_WP, P_mech_WP,"o",markerfacecolor="white", markeredgecolor="black", markersize=7)
        ax_power.spines["right"].set_position(("outward",0))
        ax_power.spines["left"].set_color("none")
        ax_power.spines["top"].set_color("none")
        ax_power.spines["bottom"].set_color("none")
        ax_power.yaxis.label.set_color("black")
        ax_power.set_ylabel("power (W)")
        ax_power.set_ylim(0,)

        ax_eta=host.twinx()
        ax_eta.patch.set_facecolor('white')
        ax_eta.plot(self.M*1000.0,self.eta*100.0,color="green")
        ax_eta.plot(self.M_meff*1000.0,self.eta_max*100.0,"d",color="red")
        ax_eta.spines["right"].set_position(("outward",120))
        ax_eta.spines["right"].set_color("green")
        ax_eta.spines["left"].set_color("none")
        ax_eta.spines["top"].set_color("none")
        ax_eta.spines["bottom"].set_color("none")
        ax_eta.yaxis.label.set_color("green")
        ax_eta.tick_params(axis="y", colors="green")
        ax_eta.set_ylabel("$\eta$ (%)")
        ax_eta.set_ylim(0,100)
        
        ax_speed=host.twinx()
        ax_speed.patch.set_facecolor('white')
        ax_speed.plot(self.M*1000.0,self.n,color="blue")
        if self.n_WP != 0 and self.M_WP != 0:
            ax_speed.plot(1000.0*self.M_WP,self.n_WP,"o",markerfacecolor="white",markeredgecolor="blue",markersize=7)
        ax_speed.spines["right"].set_position(("outward",60))
        ax_speed.spines["right"].set_color("blue")
        ax_speed.spines["left"].set_color("none")
        ax_speed.spines["top"].set_color("none")
        ax_speed.spines["bottom"].set_color("none")
        ax_speed.yaxis.label.set_color("blue")
        ax_speed.set_ylabel("speed (rpm)")
        ax_speed.yaxis.label.set_color("blue")
        ax_speed.tick_params(axis="y", colors="blue")
        ax_speed.set_ylim(0,)
        ax_speed.annotate('U={:0.1f}V'.format(self.U_N),(self.M[80]*1000*1.01,self.n[80]*1.01), color='blue', fontweight='bold')
        tstr = 'motor: {}'.format(self.motor_name)

        if addVoltagesSpeed:
            for voltage in addVoltagesSpeed:
                m_tmp = CDCMotor(U_N=voltage, R=self.R, I_0=self.I_0, k_M=self.k_M)
                ax_speed.plot(m_tmp.M*1000.0,m_tmp.n,color="blue", linestyle=":")
                ax_speed.annotate('U={:0.1f}V'.format(m_tmp.U_N),(m_tmp.M[80]*1000*1.01,m_tmp.n[80]*1.01), color='blue', fontweight='normal')
            print('voltage constant = {:0.0f}RPM/V'.format((self.n_0-m_tmp.n_0)/(self.U_N-m_tmp.U_N)))


        th=plt.text(0.65,0.95, tstr,
                    horizontalalignment='left',
                    verticalalignment='center',
                    transform=host.transAxes,
                    bbox=dict(fc='white',ec='black')
                    )


        plt.show()

        

class CDCMotorWithGearbox(CDCMotor):
    def __init__(self,**kwargs):
        super(CDCMotorWithGearbox, self).__init__(**kwargs)
        self.M_WP_system = 0
        self.n_WP_system = 0
        self.GB_name = ''
        self.set_key_values_system(**kwargs)

    def set_key_values_system(self,**kwargs):
        for key, value in kwargs.items():
            if key=="GB_ratio" : self.GB_ratio = value
            if key=="GB_eta" : self.GB_eta = value
            if key=="M_WP_system" : self.M_WP_system = value
            if key=="n_WP_system" : self.n_WP_system = value
            if key=="GB_name" : self.GB_name = value
        self.calc_system_values()

    def calc_system_values(self):
        self.M_system = self.M*self.GB_ratio*self.GB_eta
        self.M_0_system = self.M_0*self.GB_ratio*self.GB_eta
        self.M_maxpower_system = self.M_maxpower*self.GB_ratio*self.GB_eta
        self.M_S_system = self.M_S*self.GB_ratio*self.GB_eta
        self.M_meff_system = self.M_meff*self.GB_ratio*self.GB_eta
    
        self.n_system = self.n/self.GB_ratio
        self.n_0_system = self.n_system[0]
        self.n_meff_system = self.calc_n_from_M(self.M_meff)/self.GB_ratio   
        self.n_maxpower_system = self.calc_n_from_M(self.M_maxpower)/self.GB_ratio  

        self.P_mech_system = self.M_system * self.n_system *np.pi/30
        self.P_meff_system = self.M_meff_system * self.n_meff_system *np.pi/30
        self.P_maxpower_system = self.M_maxpower_system * self.n_maxpower_system *np.pi/30
        self.eta_system = self.eta*self.GB_eta
        self.eta_max_system = self.eta_max*self.GB_eta
        self.eta_maxpower_system = self.calc_eta_from_M(self.M_maxpower)*self.GB_eta

    def tune_voltage_to_working_point(self):
        self.n_WP=self.n_WP_system*self.GB_ratio
        self.M_WP=self.M_WP_system/self.GB_ratio/self.GB_eta
        super().tune_voltage_to_working_point()
        self.calc_system_values()

    def print_parameter(self):
        super().print_parameter()
        print('')
        print('gearbox input parameter')
        print('parameter\tred. ratio\tgb efficiency')
        print('unit\t\t-\t\t%')
        print('value\t\t{:0.1f}:1\t\t{:0.1f}'.format(self.GB_ratio,self.GB_eta*100.0))
        print('')

        print('system performance data:')
        print('parameter\tunit\tno-load\t\t@max eff.\t@max power\tstall')
        print('speed\t\tRPM\t{:0.0f}\t\t{:0.0f}\t\t{:0.0f}\t\t{:0.0f}'.format(self.n_0_system, self.n_meff_system, self.n_maxpower_system, 0))
        print('current\t\tA\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}'.format(self.I_0*self.GB_eta\
            , self.I_meff*self.GB_eta, self.calc_I_from_M(self.M_maxpower)*self.GB_eta,self.I_S*self.GB_eta))
        print('torque\t\tNm\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}\t\t{:0.3f}'.format(self.M_0_system, self.M_meff_system, self.M_maxpower_system, self.M_S_system))
        print('power\t\tW\t{:0.2f}\t\t{:0.2f}\t\t{:0.2f}\t\t{:0.2f}'.format(0, self.P_meff_system, self.P_maxpower_system,0))
        print('eff.\t\t%\t{:0.1f}\t\t{:0.1f}\t\t{:0.1f}\t\t{:0.1f}'.format(0, self.eta_max_system*100.0,\
            self.eta_maxpower_system,0))


    def export_to_excel(self):
        row, col = super().export_to_excel()
        wb = load_workbook(self.file_name)
        ws = wb.active
        row += 2
        col = 1
        titleStr = "System performance data"
        headerStr = ['', 'unit', '@no-load', '@max. eff.', '@max. power', 'stall']
        data = [\
            ['speed', 'RPM', self.n_0_system, self.n_meff_system, self.n_maxpower_system, 0],\
            ['current', 'A', self.I_0*self.GB_eta, self.I_meff*self.GB_eta, self.calc_I_from_M(self.M_maxpower)*self.GB_eta, self.I_S*self.GB_eta],\
            ['torque', 'Nm', self.M_0_system, self.M_meff_system, self.M_maxpower_system, self.M_S_system],\
            ['power', 'W', 0, self.P_meff_system, self.P_maxpower_system,0],\
            ['eff.', '%', 0, self.eta_max_system*100.0, self.eta_maxpower_system, 0]\
            ]
        self._excel_add_table_to_worksheet(ws, row, col, titleStr, headerStr, data)
        wb.save(self.file_name)
        

    def plotSystemCurves(self):
        plt.figure(figsize=(12,8))
        host = plt.subplot(111)
        host.set_title(self.application)
        #generate host plot
        plt.subplots_adjust(right=0.75)
        plt.subplots_adjust(left=0.1)
        plt.subplots_adjust(bottom=0.10)
        host.plot(self.M_system*1000.0,self.I*self.GB_eta,color="red")
        host.plot([1000.0*self.M_meff_system,1000.0*self.M_meff_system],[0.0, 1.2*self.I_S],":",color="black")
        host.plot([1000.0*self.M_maxpower_system,1000.0*self.M_maxpower_system],[0.0, 1.2*self.I_S],":",color="black")
        if self.M_WP_system != 0:
            host.plot([1000.0*self.M_WP_system,1000.0*self.M_WP_system],[0.0, 1.2*self.I_S],"-",color="cyan")
        host.set_xlabel("torque (mNm)")
        host.set_ylabel("current (A)")
        host.yaxis.label.set_color("red")
        host.spines["left"].set_color("red")
        host.tick_params(axis="y", colors="red")
        host.grid(True)
        host.set_xlim(0,1000.0*self.M_S_system)
        host.set_ylim(0,1.2*self.I_S*self.GB_eta)

        
        ax_power=host.twinx()
        ax_power.plot(self.M_system*1000.0,self.P_mech_system,".-",color="black")
        ax_power.plot(1000.0*self.M_meff_system,self.P_meff_system,"d",color="red")
        ax_power.plot(1000.0*self.M_maxpower_system,self.P_maxpower_system,"d",color="red")
        if self.M_WP_system != 0 and self.n_WP_system != 0:
            P_mech_WP_system = self.M_WP_system * self.n_WP_system * np.pi / 30.0
            ax_power.plot(1000.0*self.M_WP_system, P_mech_WP_system,"o",markerfacecolor="white", markeredgecolor="black", markersize=7)
        ax_power.spines["right"].set_position(("outward",0))
        ax_power.spines["left"].set_color("none")
        ax_power.spines["top"].set_color("none")
        ax_power.spines["bottom"].set_color("none")
        ax_power.yaxis.label.set_color("black")
        ax_power.set_ylabel("power (W)")
        ax_power.set_ylim(0,)

        ax_eta=host.twinx()
        ax_eta.plot(self.M_system*1000.0,self.eta_system*100.0,color="green")
        ax_eta.plot(self.M_meff_system*1000.0,self.eta_max_system*100.0,"d",color="red")
        ax_eta.spines["right"].set_position(("outward",120))
        ax_eta.spines["right"].set_color("green")
        ax_eta.spines["left"].set_color("none")
        ax_eta.spines["top"].set_color("none")
        ax_eta.spines["bottom"].set_color("none")
        ax_eta.yaxis.label.set_color("green")
        ax_eta.tick_params(axis="y", colors="green")
        ax_eta.set_ylabel("$\eta$ (%)")
        ax_eta.set_ylim(0,100)
        
        ax_speed=host.twinx()
        ax_speed.plot(self.M_system*1000.0,self.n_system,color="blue")
        if self.n_WP_system != 0 and self.M_WP_system != 0:
            ax_speed.plot(1000.0*self.M_WP_system,self.n_WP_system,"o",markerfacecolor="white",markeredgecolor="blue",markersize=7)
        ax_speed.spines["right"].set_position(("outward",60))
        ax_speed.spines["right"].set_color("blue")
        ax_speed.spines["left"].set_color("none")
        ax_speed.spines["top"].set_color("none")
        ax_speed.spines["bottom"].set_color("none")
        ax_speed.yaxis.label.set_color("blue")
        ax_speed.set_ylabel("speed (rpm)")
        ax_speed.yaxis.label.set_color("blue")
        ax_speed.tick_params(axis="y", colors="blue")
        ax_speed.set_ylim(0,)
        tstr = 'motor: {}\ngearbox: {}'.format(self.motor_name, self.GB_name)

        th=plt.text(0.65,0.95, tstr,
                    horizontalalignment='left',
                    verticalalignment='center',
                    transform=host.transAxes,
                    bbox=dict(fc='white',ec='black')
                    )
        plt.show()
        

        


if __name__ == "__main__":
    # R=1.44 (-25°C)
    # R=2.00 (125°C)
    
    c = 4.7
    U_N = 10
    T = -40
    m_T = (2.0-1.44)/(125+25)
    T_0 = 1.44 - m_T*(-25)
    R = m_T*T+T_0*c
    I_S = U_N / R
    k_M = 0.005985*np.sqrt(c)
    M_S = 0.017099
    I_0 = 0.13*6/U_N
    n_WP = 300
    M_WP = 0.005
    dcmotor=CDCMotor(U_N=U_N, I_0=I_0, k_M=k_M, R=R, H=0.61, Theta=6.7, n_WP=n_WP, M_WP=M_WP, application="166_A / LiDAR", motor_name="BO2015_Version 10V")
    dcmotor.print_parameter()
    dcmotor.tune_voltage_to_working_point()
    # dcmotor.print_parameter()
    # dcmotor.plotCurves(addVoltagesSpeed=[3,4,5,6])
    # dcmotor.plotCurves()
    dcmotor.list_spec_table()
 
    # GB_eta=0.8
    # GB_ratio=100
    # GB_name='100:1'
    # gbmotor=CDCMotorWithGearbox(U_N=U_N, I_0=I_0, k_M=k_M, R=R, GB_ratio=GB_ratio, GB_eta=GB_eta, \
    #     n_WP_system=n_WP/GB_ratio, M_WP_system=M_WP*GB_ratio*GB_eta, \
    #     application="Test with GB", motor_name="BO2015_Version 10V", GB_name=GB_name)
    # gbmotor.print_parameter()
    # gbmotor.tune_voltage_to_working_point()
    # gbmotor.print_parameter()
    # gbmotor.plotSystemCurves()